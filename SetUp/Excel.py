from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time
from SetUp.Get_dBm_Meas import Get_dBm_Meas
from SetUp.NF_Functions import NF


class Excel:
    def __init__(self):
        print('working on excel')
        self.Header = {'Made': 'A',
                       'Path': 'B',
                       'temp(C)': 'C',
                       'Gain': 'D',
                       'Min': 'E',
                       'Max': 'F',
                       'Offset (dBm)': 'G',
                       'Model': 'H',
                       'Time Stamp': 'I',
                       'project': 'J',
                       'Pass/Fail': 'K'}

    def Create_WorkBook_wHeader(self,file_path,file_name):
        wb = Workbook()
        ws = wb.active

        ws[self.Header['Made'] + str(1)] = 'Made'
        ws[self.Header['Path']+ str(1)] = 'Path'
        ws[self.Header['temp(C)'] + str(1)] = 'temp(C)'
        ws[self.Header['Gain'] + str(1)] = 'Gain'
        ws[self.Header['Min']+ str(1)] = 'Min'
        ws[self.Header['Max'] + str(1)] = 'Max'
        ws[self.Header['Offset (dBm)'] + str(1)] = 'Offset (dBm)'
        ws[self.Header['Model']+ str(1)] = 'Model'
        ws[self.Header['Time Stamp'] + str(1)] = 'Time Stamp'
        ws[self.Header['project'] + str(1)] = 'project'
        ws[self.Header['Pass/Fail'] + str(1)] = 'Pass/Fail'
        wb.save(file_path + file_name)

    def Create_Gain_Meas_Workbook(self,file_path,file_name):
        wb = Workbook()
        ws = wb.active


        wb.save(file_path + file_name)
    #def Add_Data_Gain_Meas(self,file_path,Made,Pipe,Data):


    def Add_data(self,file_path,file_name,Data):
        wb = load_workbook(file_path+file_name)
        sheet = wb.get_sheet_by_name("Sheet")

        current_col = sheet.max_row + 1

        sheet[self.Header['clock_name'] + str(current_col)] = Data['clock_name']
        sheet[ self.Header['clock_freq']+ str(current_col)] = Data['clock_freq']
        sheet[self.Header['pipe'] + str(current_col)] = Data['pipe']
        sheet[self.Header['temp(C)'] + str(current_col)] = Data['temperature']
        sheet[self.Header['test case']+ str(current_col)] = Data['test_case']
        sheet[self.Header['#'] + str(current_col)] = Data['#']
        sheet[self.Header['offset_freq (MHz)'] + str(current_col)] = Data['offset_freq']
        sheet[self.Header['value (dBC)']+ str(current_col)] = Data['value']
        sheet[self.Header['Time Stamp'] + str(current_col)] = Data['Time_Stamp']
        sheet[self.Header['project'] + str(current_col)] = Data['project']


#        sheet.append(Data)

        wb.save(file_path + file_name)

    def Add_Gain_data(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp,cen_freq,g_file_name,num, data_set):

        Data = Get_dBm_Meas()
        Gain = Data.add_Gain_Meas(gain_file_path+g_file_name,cen_freq,g_file_name)
        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4
        row = made * 4 + r + 1
        ws["C{}".format(str(row))] = Gain[2] # measured gain
        ws["J{}".format(str(row))] = Gain[1] # measured freq
        ws["D{}".format(str(row))] = power  # amplitude
        ws["E{}".format(str(row))] = offset  # offset


        wb.save(excel_file_path)

    def Add_ISO_data(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp, peak_ind,iso_made,iso_pipe,cen_freq):

        Data = Get_dBm_Meas()
        Gain = Data.add_Iso_Meas(gain_file_path, peak_ind,cen_freq)

        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4

        if iso_pipe == 50:
            i = 1
        elif iso_pipe == 52:
            i = 2
        elif iso_pipe == 54:
            i = 3
        elif iso_pipe == 56:
            i = 4
        row1 = made * 4 + r + 1
        column1 = iso_made*4 + i + 2

        g = ws.cell(row=row1, column=column1)
        g.value=Gain[1]
        #if '50.txt' in gain_file_path:
         #   Gain[1] = ws.cell(row=row1,column=column1)

        #elif '52.txt' in gain_file_path:
         #   ws["E{}".format(str(row))] = Gain[1]
          #  ws["F{}".format(str(row))] = Gain[0]
        #elif '54.txt' in gain_file_path:
         #   ws["G{}".format(str(row))] = Gain[1]
          #  ws["H{}".format(str(row))] = Gain[0]
        #elif '56.txt' in gain_file_path:
         #   ws["I{}".format(str(row))] = Gain[1]
          #  ws["J{}".format(str(row))] = Gain[0]

        #if'isofrom' in gain_file_path:
         #   print('ISO PATH')
        #else:
         #   ws["K{}".format(str(row))] = power  # amplitude
          #  ws["L{}".format(str(row))] = offset  # offset



        wb.save(excel_file_path)

    def Add_Flat_Data(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp,cen_freq,g_file_name,num,data_set):
        Data = Get_dBm_Meas()
        Gain = Data.add_Gain_Meas(gain_file_path,cen_freq,g_file_name)
        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4
        row = made * 4 + r + 1
        if '3740.txt' in gain_file_path or '3460' in gain_file_path:
            ws["C{}".format(str(row))] = Gain[2]
            ws["D{}".format(str(row))] = Gain[1]
        if '3840.txt' in gain_file_path or '3500' in gain_file_path:
            ws["E{}".format(str(row))] = Gain[2]
            ws["F{}".format(str(row))] = Gain[1]
        if '3940.txt' in gain_file_path or '3540' in gain_file_path:
            ws["G{}".format(str(row))] = Gain[2]
            ws["H{}".format(str(row))] = Gain[1]
        ws["I{}".format(str(row))] = power  # amplitude
        ws["J{}".format(str(row))] = offset  # offset
        wb.save(excel_file_path)

    def Add_IIP3_Data(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp, dist,data_set, num_cap):
        Data = Get_dBm_Meas()
        #Gain = Data.add_IP3_Meas(gain_file_path, dist, gain_file_path, made,pipe)
        Gain = Data.add_IP3_Meas_AVG(gain_file_path, dist, gain_file_path, made, pipe, data_set, num_cap)
        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4
        row = made * 4 + r + 1

        ws["C{}".format(str(row))] = Gain[4]
        ws["D{}".format(str(row))] = Gain[3]/1e6
        ws["E{}".format(str(row))] = Gain[6]
        ws["F{}".format(str(row))] = Gain[5]/1e6
        ws["G{}".format(str(row))] = Gain[8]
        ws["H{}".format(str(row))] = Gain[7]/1e6
        ws["I{}".format(str(row))] = Gain[10]
        ws["J{}".format(str(row))] = Gain[9]/1e6

        ws["K{}".format(str(row))] = power  # amplitude
        ws["L{}".format(str(row))] = offset  # offset
        wb.save(excel_file_path)
    def Add_IIP3_Data_cont(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp, dist,data_set, num_cap,z):
        Data = Get_dBm_Meas()
        #Gain = Data.add_IP3_Meas(gain_file_path, dist, gain_file_path, made,pipe)
        Gain = Data.add_IP3_Meas_AVG(gain_file_path, dist, gain_file_path, made, pipe, data_set, num_cap)
        wb = load_workbook(excel_file_path)
        if pipe == 50:
            ws = wb.worksheets[0]
        elif pipe == 52:
            ws = wb.worksheets[1]
        elif pipe == 54:
            ws = wb.worksheets[2]
        elif pipe == 56:
            ws = wb.worksheets[3]

        r = z+1
        row = r

        ws["C{}".format(str(row))] = Gain[4]
        ws["D{}".format(str(row))] = Gain[3]/1e6
        ws["E{}".format(str(row))] = Gain[6]
        ws["F{}".format(str(row))] = Gain[5]/1e6
        ws["G{}".format(str(row))] = Gain[8]
        ws["H{}".format(str(row))] = Gain[7]/1e6
        ws["I{}".format(str(row))] = Gain[10]
        ws["J{}".format(str(row))] = Gain[9]/1e6

        ws["K{}".format(str(row))] = power  # amplitude
        ws["L{}".format(str(row))] = offset  # offset
        wb.save(excel_file_path)
        if Gain[2] < power+24:
            g_pass = False
            print('Gain too low')
            if Gain[2] > -23:
                g_pass = True
        else:
            g_pass = True
        return g_pass
    def Add_NF_data(self,excel_file_path,pipe, made, gain_file_path, power, offset, temp, data_avg, num_cap, BW,Gain,freq):

        nfPlot = NF()
        nf = nfPlot.plot_NF(data_avg, num_cap, BW, Gain, pipe,made,gain_file_path,freq)
        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4
        row = made * 4 + r + 1
        ws["C{}".format(str(row))] = Gain # measured gain
        ws["D{}".format(str(row))] = power  # amplitude
        ws["E{}".format(str(row))] = offset  # offset
        ws["G{}".format(str(row))] = num_cap  # number of captures
        ws["H{}".format(str(row))] = BW/(1e3)  # BW
        ws["I{}".format(str(row))] = nf[1] #Output noise sum
        ws["J{}".format(str(row))] = nf[2]  # Thermal Noise
        ws["K{}".format(str(row))] = nf[3]  # Noise Factor



        wb.save(excel_file_path)

    def Add_P1dB_data(self,excel_file_path,pipe, made, gain_file_path, powIn, offset, temp,cen_freq,g_file_name,num,data_set):

        Data = Get_dBm_Meas()
        Gain = Data.add_Gain_Meas(gain_file_path,cen_freq,g_file_name)

        wb = load_workbook(excel_file_path)
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4

        row1 = (made * 4) + r + 1

        column1 = (powIn + 40) + 5


        pow2 = ws.cell(row = row1,column = column1)
        pow2.value=Gain[2]


        pOut = Gain[2]



        wb.save(excel_file_path)
        return pOut

    def Add_P1dB_Point(self, excel_file_path, pipe, made, temp, p1):


        wb = load_workbook(excel_file_path)
        ws = wb.worksheets[0]
        if temp == 25:
            ws = wb.worksheets[0]
        elif temp == -40:
            ws = wb.worksheets[1]
        elif temp == 95:
            ws = wb.worksheets[2]

        if pipe == 50:
            r = 1
        elif pipe == 52:
            r = 2
        elif pipe == 54:
            r = 3
        elif pipe == 56:
            r = 4

        row1 = (made * 4) + r + 1

        column1 = 4

        p1db = ws.cell(row=row1, column=4)
        p1db.value = p1

        wb.save(excel_file_path)
    def Add_ATT_data(self,excel_file_path,pipe, made, gain_file_path, power,cen_freq,g_file_name,ATT,row,PowValue):

        Data = Get_dBm_Meas()
        Gain = Data.add_Gain_Meas(g_file_name,cen_freq,g_file_name)
        wb = load_workbook(excel_file_path)
        if made == 0:
            ws = wb.worksheets[0]
        else:
            ws = wb.worksheets[1]



        if pipe == 50:
            ws["C{}".format(str(row))] = power
            ws["D{}".format(str(row))] = Gain[2]  # measured power
            ws["E{}".format(str(row))] = ATT
            ws["G{}".format(str(row))] = PowValue
            ws["I{}".format(str(row))] = Gain[3]
            ws["I{}".format(str(row))] = Gain[4]

        elif pipe == 52:
            ws["M{}".format(str(row))] = power
            ws["N{}".format(str(row))] = Gain[2]  # measured power
            ws["O{}".format(str(row))] = ATT
            ws["Q{}".format(str(row))] = PowValue
            ws["S{}".format(str(row))] = Gain[3]
            ws["T{}".format(str(row))] = Gain[4]
        elif pipe == 54:
            ws["W{}".format(str(row))] = power
            ws["X{}".format(str(row))] = Gain[2]  # measured power
            ws["Y{}".format(str(row))] = ATT
            ws["AA{}".format(str(row))] = PowValue
            ws["AC{}".format(str(row))] = Gain[3]
            ws["AD{}".format(str(row))] = Gain[4]
        elif pipe == 56:
            ws["AG{}".format(str(row))] = power
            ws["AH{}".format(str(row))] = Gain[2]  # measured power
            ws["AI{}".format(str(row))] = ATT
            ws["AK{}".format(str(row))] = PowValue
            ws["AM{}".format(str(row))] = Gain[3]
            ws["AN{}".format(str(row))] = Gain[4]






        wb.save(excel_file_path)



if __name__ == "__main__":
    start_time = time.time()
    excel = Excel()
    excel.Create_WorkBook_wHeader("C:\\Users\\eryoung\\Desktop\\PycharmProjects\\Testing\\Code\\Data\\","trial2.xlsx")


    Data = {'clock_name': "some clock"}
    Data['clock_freq'] = 2565498
    Data['pipe'] = 1
    Data['temperature'] = 25
    Data['test_case'] = "some clock test"
    Data['#'] = 0
    Data['offset_freq'] = 102355
    Data['value'] = 1.23
    Data['Time_Stamp'] = 14523
    i = 0
    while(i<100):
        excel.Add_data("C:\\Users\\eryoung\\Desktop\\PycharmProjects\\Testing\\Code\\Data\\","trial2.xlsx",Data)
        i = i+1

    print("time taken in seconds: %f" %(time.time()-start_time))

    #takes about 2 minutes to write 1000 lines